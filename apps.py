import streamlit as st
from openpyxl import load_workbook
import datetime
import random
import pandas as pd
import os
from PIL import Image
import time
import plotly.graph_objects as go

from generating import create_label, create_label_with_manual_issues, initialize_product
from exporting import create_pdf
from utilities import get_device_id, get_batch_id
from ocr_validation import QR_Image_Validation

st.set_page_config(page_title="Smart Product Labeling", layout="centered")
initialize_product()

TAB1, TAB2 = st.tabs(["Entry Module", " Admin - Traceability"])

# ------------------- INPUT MODULE ---------------------
with TAB1:
    st.header("Product Entry System")
    mode = st.radio("Entry Mode", ["Admin", "Sensor"])

    if mode == "Admin":
        st.subheader(" Admin-Controlled Product Entry")
        device_id = st.text_input("Device ID", value=get_device_id())
        batch_id = st.text_input("Batch ID", value=get_batch_id())
        mfg_date = st.date_input("Manufacturing Date", datetime.date.today())
        rohs = st.radio("RoHS Compliant?", ["Yes", "No"])
        tm1_issue = st.text_input("TM1 Issue Comment", placeholder="e.g., Barcode scanner delay")
        tm2_issue = st.text_input("TM2 Issue Comment", placeholder="e.g., Label smudged")
        tm3_issue = st.text_input("TM3 Issue Comment", placeholder="e.g., Sensor glitch")
        admin_comment = st.text_area("Admin Comment (optional)")

        if st.button("Submit"):
            if device_id and batch_id:
                create_label_with_manual_issues(
                    device_id, batch_id, str(mfg_date), rohs,
                    entry_mode="Admin",
                    comment=admin_comment,
                    manual_issues=[tm1_issue, tm2_issue, tm3_issue]
                )
                st.success(" Product tested and recorded with issue comments.")
                qr_path = f"qr_codes/{device_id}.png"
                if os.path.exists(qr_path):
                    st.image(qr_path, caption=" Generated QR Code", width=200)
            else:
                st.warning("Please fill in all required fields.")

    else:
        st.subheader("Simulated Sensor Entry")
        if st.button("Auto-Generate Sensor Entry"):
            device_id = get_device_id()
            batch_id = get_batch_id()
            mfg_date = datetime.date.today()
            rohs = random.choice(["Yes", "No"])

            create_label(device_id, batch_id, str(mfg_date), rohs, "Sensor", "")
            st.success(f"Sensor entry submitted: {device_id} | {batch_id}")
            qr_path = f"qr_codes/{device_id}.png"
            if os.path.exists(qr_path):
                st.image(qr_path, caption=" Generated QR Code", width=200)

# ------------------- TRACKING MODULE ---------------------
with TAB2:
    st.header("Admin Access")
    password = st.text_input("Enter Admin Password", type="password")

    if password == "admin123":
        st.success("Access Granted")
        device_id_search = st.text_input("Search Device ID")

        wb = load_workbook("inspection_log.xlsx")
        ws = wb.active
        data = list(ws.iter_rows(values_only=True))
        headers = [h.strip() if isinstance(h, str) else h for h in data[0]]
        rows = data[1:]

        if device_id_search:
            filtered = [dict(zip(headers, r)) for r in rows if r[0] == device_id_search]

            if filtered:
                st.subheader(f"Report for Device: {device_id_search}")
                df = pd.DataFrame(filtered)
                df.columns = df.columns.str.strip()
                st.dataframe(df)

                latest = filtered[-1]
                st.markdown(f"""
                - **Batch ID**: `{latest.get('Batch ID', 'N/A')}`
                - **RoHS**: `{latest.get('RoHS', 'N/A')}`
                - **Entry Mode**: `{latest.get('Entry Mode', 'N/A')}`
                - **Admin Comment**: `{latest.get('Admin Comment', 'None')}`

                ###  Testing Summary:
                - **TM1**: `{latest.get('TM1 Status', 'N/A')}` | {latest.get('TM1 Time (s)', 'N/A')}s | {latest.get('TM1 Issue', 'No issue')}`
                - **TM2**: `{latest.get('TM2 Status', 'N/A')}` | {latest.get('TM2 Time (s)', 'N/A')}s | {latest.get('TM2 Issue', 'No issue')}`
                - **TM3**: `{latest.get('TM3 Status', 'N/A')}` | {latest.get('TM3 Time (s)', 'N/A')}s | {latest.get('TM3 Issue', 'No issue')}`
                """)

                st.markdown("### Download Device Traceability Report")
                if st.button("Generate PDF Report"):
                    try:
                        qr_path = f"qr_codes/{latest.get('Device ID')}.png"
                        pdf_path = f"{latest.get('Device ID')}_trace_report.pdf"
                        create_pdf(latest, qr_path, pdf_path)
                        with open(pdf_path, "rb") as f:
                            st.download_button(" Download PDF", f, file_name=pdf_path)
                    except Exception as e:
                        st.error(f"Failed to generate PDF: {e}")
            else:
                st.warning("Device ID not found.")

        # ---------------- QR IMAGE VALIDATION ----------------
        st.markdown("---")
        st.subheader("Validate Printed Label (via OCR)")

        uploaded_img = st.file_uploader("Upload Scanned QR Label Image", type=["png", "jpg", "jpeg"])
        if uploaded_img:
            try:
                image = Image.open(uploaded_img).convert("RGB")
                temp_path = "temp_uploaded_label.png"
                image.save(temp_path)

                st.image(temp_path, caption="📷 Uploaded QR Label", width=300)
                result = QR_Image_Validation(temp_path)

                if result["device_id"] != "N/A" and result["batch_id"] != "N/A":
                    st.success(f"Extracted → Device ID: `{result['device_id']}`, Batch ID: `{result['batch_id']}`")
                    st.info(f"Time Taken: {result['time']}s")
                    st.markdown(f"**QR Decode Status:** `{result['status']}`")
                    st.markdown(f"**Issue:** `{result['issue']}`")
                    st.code(result["ocr_text"], language="text")

                    matched = [
                        dict(zip(headers, r))
                        for r in rows
                        if r[0] == result["device_id"] and r[1] == result["batch_id"]
                    ]

                    if matched:
                        st.markdown("###  Matched Log Entry:")
                        st.dataframe(pd.DataFrame(matched))

                        # Generate pdf report
                        latest_match = matched[-1]
                        st.markdown("### Download Traceability PDF from QR Match")
                        if st.button("Generate PDF from QR"):
                            try:
                                qr_path = f"qr_codes/{latest_match.get('Device ID')}.png"
                                pdf_path = f"{latest_match.get('Device ID')}_trace_report.pdf"
                                create_pdf(latest_match, qr_path, pdf_path)
                                with open(pdf_path, "rb") as f:
                                    st.download_button("Download PDF", f, file_name=pdf_path)
                            except Exception as e:
                                st.error(f"Failed to generate PDF: {e}")
                    else:
                        st.error("Device not found in log.")
                else:
                    st.warning("Could not extract valid Device ID and Batch ID from QR code.")
            except Exception as e:
                st.error(f"QR Decode failed: {e}")

        # ---------------- TM ERROR GRAPH ----------------
        st.markdown("---")
        st.subheader("TM Failure Analytics")

        tm1_fails = sum(1 for r in rows if r[6] == "REJECT")
        tm2_fails = sum(1 for r in rows if r[10] == "REJECT")
        tm3_fails = sum(1 for r in rows if r[14] == "REJECT")

        fig = go.Figure(data=[
            go.Bar(name='TM Failures', x=["TM1", "TM2", "TM3"], y=[tm1_fails, tm2_fails, tm3_fails])
        ])
        fig.update_layout(
            title="Failure Count per Test Machine",
            xaxis_title="Test Machine",
            yaxis_title="Number of Failures",
            height=400
        )
        st.plotly_chart(fig, use_container_width=True)

    elif password:
        st.error("Incorrect Password")
