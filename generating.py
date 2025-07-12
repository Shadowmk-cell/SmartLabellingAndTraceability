# generating the label
import qrcode
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import time
from utilities import get_machine_testing_time

QR_FOLDER = "qr_codes"
EXCEL_FILE = "inspection_log.xlsx"

def initialize_product():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Inspection Log"
        ws.append([
            "Device ID", "Batch ID", "RoHS", "QR Code Path", "Entry Mode", "Admin Comment",
            "TM1 Status", "TM1 Time (s)", "TM1 Timestamp", "TM1 Issue",
            "TM2 Status", "TM2 Time (s)", "TM2 Timestamp", "TM2 Issue",
            "TM3 Status", "TM3 Time (s)", "TM3 Timestamp", "TM3 Issue"
        ])
        wb.save(EXCEL_FILE)

def create_label(device_id, batch_id, mfg_date, rohs, entry_mode="Sensor", comment=""):
    initialize_product()
    os.makedirs(QR_FOLDER, exist_ok=True)

    qr_data = f"Device ID: {device_id}\nBatch ID: {batch_id}\nDate: {mfg_date}\nRoHS: {rohs}"
    qr_path = os.path.join(QR_FOLDER, f"{device_id}.png")
    qrcode.make(qr_data).save(qr_path)

    machine_data = []
    for machine in ["TM1", "TM2", "TM3"]:
        time_taken = get_machine_testing_time(machine)
        time.sleep(time_taken)
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        status = "PASS" if time_taken <= 5.0 else "REJECT"
        issue_comment = "Entry Delay" if time_taken > 5.0 else "No issue"
        machine_data.extend([status, time_taken, timestamp, issue_comment])

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        device_id, batch_id, rohs, qr_path, entry_mode, comment,
        *machine_data
    ])
    wb.save(EXCEL_FILE)
    print(f" Sensor entry logged: {device_id}")

def create_label_with_manual_issues(device_id, batch_id, mfg_date, rohs, entry_mode="Admin", comment="", manual_issues=None):
    initialize_product()
    os.makedirs(QR_FOLDER, exist_ok=True)

    qr_data = f"Device ID: {device_id}\nBatch ID: {batch_id}\nDate: {mfg_date}\nRoHS: {rohs}"
    qr_path = os.path.join(QR_FOLDER, f"{device_id}.png")
    qrcode.make(qr_data).save(qr_path)

    machine_data = []
    machines = ["TM1", "TM2", "TM3"]

    for idx, machine in enumerate(machines):
        time_taken = get_machine_testing_time(machine)
        time.sleep(time_taken)
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        status = "PASS" if time_taken <= 5.0 else "REJECT"
        issue_comment = (
            manual_issues[idx] if manual_issues and idx < len(manual_issues) and manual_issues[idx]
            else ("Entry Delay" if time_taken > 5.0 else "No issue")
        )
        machine_data.extend([status, time_taken, timestamp, issue_comment])

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        device_id, batch_id, rohs, qr_path, entry_mode, comment,
        *machine_data
    ])
    wb.save(EXCEL_FILE)
    print(f" Admin entry logged with manual issues: {device_id}")
